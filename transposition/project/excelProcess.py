import matplotlib.pyplot as plt
import xlrd
import openpyxl
import numpy as np

origin_list = [0.0979,0.0981,0.0937,0.1111,0.3613,0.1234,0.1064,0.1133,0.1028,0.4222,0.398,0.154\
,0.1292,0.1601,0.1505,0.1771,0.5589,0.19,0.3863,0.4156,0.4467,0.2076,0.4283,0.2677\
,0.1533,0.1206,0.1643,0.1449,0.6063,0.1527,0.1265,0.6844,0.1856,0.209,0.2184,0.2797\
,0.4113,0.1744,0.6115,0.199,0.1992,0.1616,0.1803,0.1769,0.5697,0.1937,0.157,0.5558\
,0.1302,0.1755,0.1628,0.1326,0.1406,0.1352,0.113,0.1132,0.1841,0.1467,0.1447,0.2954\
,0.1506,0.1926,0.1481,0.5185,0.2137,0.1544,0.1423,0.1494,0.2039,0.3855,0.258,0.1779\
,0.1129,0.1359,0.1542,0.1348,0.518,0.1442,0.0974,0.1108,0.1257,0.3784,0.1783,0.1841\
,0.3895,0.6134,0.161,0.1483,0.515,0.181,0.1479,0.1296,0.1771,0.2064,0.1583,0.1109]
sampleDict = dict(zip(origin_list,range(1,len(origin_list)+1)))
sort_ori_list = sorted(origin_list)

effect_list = [1-x/max(origin_list) for x in origin_list]
sort_eff_list = sorted(effect_list)

def drawPics(ori,sort_ori):
    good_with_value = list()
    sort_good_list = sort_ori[-10:]
    for elem in sort_good_list:
        good_with_value.append(sampleDict[elem])
    #原始点图+sorted点图
    plt.figure(figsize=(6,6))

    plt.figure(1)
    ax1 = plt.subplot(211)
    plt.scatter(range(1,len(ori)+1),ori,10, color = 'red')
    plt.scatter(good_with_value,sort_good_list,20,color = 'blue')
    plt.xlim((0,100))
    plt.ylim((0,0.8))
    x_ticks = np.linspace(0, 100, 11)
    plt.xticks(x_ticks)
    x1 = range(1,101)
    y1 = sort_good_list[0]-0.05
    plt.fill_between(x1,y1,y2 = 0.8,facecolor='purple',alpha = 0.3)
    aver_ori_list = [sum(ori)/len(ori)]*100
    plt.plot(range(1,len(aver_ori_list)+1),aver_ori_list,'--')
    
    ax2 = plt.subplot(212)
    plt.xlim((0,100))
    plt.ylim((0,1))
    x_ticks = np.linspace(0, 100, 11)
    plt.xticks(x_ticks)
    plt.scatter(range(1,len(sort_ori)+1),sort_ori,10,color='red')    
    plt.scatter(range(len(sort_ori)-len(sort_good_list)+1,len(sort_ori)+1),sort_good_list,10,color='blue')
    x2 = range(len(ori)-10,len(ori)+2)
    plt.fill_between(x2,1,facecolor='purple',alpha = 0.3)
    plt.show()

def print_data_to_excel(ori,sort_ori,eff_ori,name):

    wb = openpyxl.load_workbook(source_path+name)
    sheet = wb.create_sheet()
    # sheet = wb.worksheets[sheet_num]
    
    sheet.cell(1,1).value = 'No.'
    sheet.cell(1,2).value = 'OD600'
    sheet.cell(1,3).value = 'Sorted Res'
    sheet.cell(1,4).value = 'Effect Res'

    
    for i,v in enumerate(sort_ori):
        sheet.cell(i+2,1).value = i+1
        sheet.cell(i+2,2).value = ori[i]
        sheet.cell(i+2,3).value = sort_ori[i]
        sheet.cell(i+2,4).value = eff_ori[i]
        
    
    wb.save(source_path+name)

source_path = 'E:/VSCode/code/bp3/'
if __name__ == "__main__":
    # print_data_to_excel(origin_list,sort_ori_list,sort_eff_list,'test.xlsx')
    drawPics(origin_list,sort_ori_list)
    #drawPics(effect_list,sort_eff_list)

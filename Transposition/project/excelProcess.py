import matplotlib.pyplot as plt
import xlrd
import openpyxl
import numpy as np

origin_list = [0.7525,0.3505,0.3161,0.325,0.2779,0.3011,0.3197,0.2645,0.4398,0.2736,0.2048,0.2817\
,0.3037,0.3141,0.3519,0.3769,0.3265,1.0202,0.306,0.2569,0.2439,0.2863,0.2287,0.1716\
,0.3362,0.3611,0.3466,1.0457,0.9707,1.0091,0.8536,0.963,0.3258,0.2921,0.2982,0.6359\
,0.3612,0.4224,0.3816,1.0114,1.0122,1.0518,1.0112,1.035,0.965,0.3688,0.333,0.5294\
,0.393,0.4539,1.045,0.986,1.0487,1.0112,1.0456,1.0473,1.0146,1.0108,0.4997,0.709\
,0.6247,0.4552,0.5618,0.9474,1.0354,1.0489,1.0715,1.0822,1.0317,0.4917,0.3417,1.2095\
,0.6468,0.3759,0.4164,0.4438,0.595,1.0438,0.6449,1.0679,0.5874,0.4416,0.5767,1.0206\
,0.7683,0.803,0.8024,0.6408,0.7188,0.8813,1.0662,1.1104,0.9359,0.8247,0.929,1.3497]
ori_sampleDict = dict(zip(origin_list,range(1,len(origin_list)+1)))
sort_ori_list = sorted(origin_list)

effect_list = [1-x/max(origin_list) for x in origin_list]
eff_sampleDict = dict(zip(effect_list,range(1,len(effect_list)+1)))
sort_eff_list = sorted(effect_list)

#原始点图+sorted点图
def drawPics(ori,sort_ori,sampleDict,ori_or_eff):
    good_with_value = list()
    if ori_or_eff == 1:
        sort_good_list = sort_ori[:11]
    else:
        sort_good_list = sort_ori[-10:]
    for elem in sort_good_list:
        good_with_value.append(sampleDict[elem])
    #判断图类型
    if ori_or_eff == 1:
        YLIM = 0.8
        DISTC = 9
        YLABEL = '$OD_{600}$'
        FILL_X1 = 0.6
        FILL_X2 = len(sort_good_list)+0.4
        PLOT_X1 = 1
        PLOT_X2 = len(sort_good_list)+1
        
    else:
        YLIM = 1.0
        DISTC = 11
        YLABEL = 'Effect'
        
        FILL_X1 = len(ori)-9.5
        FILL_X2 = len(ori)+2
        PLOT_X1 = len(sort_ori)-len(sort_good_list)+1
        PLOT_X2 = len(sort_ori)+1

    plt.figure(figsize=(6,8))

    plt.figure(1)
    #picture 1
    ax1 = plt.subplot(211)
    plt.scatter(range(1,len(ori)+1),ori,10, color = 'tomato')
    plt.scatter(good_with_value,sort_good_list,20,color = 'crimson')
    #共同特征
    plt.xlim((0,100))
    plt.ylim((0,YLIM))
    x_ticks = np.linspace(0, 100, 11)
    plt.xticks(x_ticks)
    y_ticks = np.linspace(0, YLIM,DISTC)
    plt.yticks(y_ticks)
    ax1.set_ylabel(YLABEL,fontsize=10)
    ax1.set_xlabel("Sequence No.",fontsize=10)

    x1 = range(-1,101)
    y1 = sort_good_list[0]-0.02
    plt.fill_between(x1,y1,y2 = sort_good_list[-1]+0.02,facecolor='c',alpha = 0.3)

    aver_ori_list = [sum(ori)/len(ori)]*105
    plt.plot(range(-1,len(aver_ori_list)-1),aver_ori_list,'--')
    #picture 2
    ax2 = plt.subplot(212)
    #共同特征
    plt.xlim((0,100))
    plt.ylim((0,YLIM))
    x_ticks = np.linspace(0, 100, 11)
    plt.xticks(x_ticks)
    y_ticks = np.linspace(0, YLIM,DISTC)
    plt.yticks(y_ticks)
    ax2.set_ylabel(YLABEL,fontsize=10)
    ax2.set_xlabel("Sequence No.",fontsize=10)

    plt.plot(range(1,len(sort_ori)+1),sort_ori, color='tomato', marker='o',linewidth=1, \
        markersize=2)
    # plt.scatter(range(1,len(sort_ori)+1),sort_ori,10,marker = '+',color='tomato',)    
    plt.plot(range(PLOT_X1,PLOT_X2),sort_good_list,marker = 'o',color='crimson',\
        linewidth=1,markersize=2)
    x2 = np.linspace(FILL_X1,FILL_X2)
    plt.fill_between(x2,1,facecolor='c',alpha = 0.3)
    plt.show()

def print_data_to_excel(ori,sort_ori,eff_ori,name):

    wb = openpyxl.load_workbook(source_path+name)
    sheet = wb.create_sheet()
    # sheet = wb.worksheets[sheet_num]
    
    sheet.cell(1,1).value = 'No.'
    sheet.cell(1,2).value = 'OD600'
    sheet.cell(1,3).value = 'Sorted Res'
    sheet.cell(1,4).value = 'Effect Res'

    
    for i,v in enumerate(sort_ori):
        sheet.cell(i+2,1).value = i+1
        sheet.cell(i+2,2).value = ori[i]
        sheet.cell(i+2,3).value = sort_ori[i]
        sheet.cell(i+2,4).value = eff_ori[i]
        
    
    wb.save(source_path+name)

source_path = 'E:/VSCode/code/bp3/'
if __name__ == "__main__":
    # print_data_to_excel(origin_list,sort_ori_list,sort_eff_list,'test.xlsx')
    drawPics(origin_list,sort_ori_list,ori_sampleDict,1)
    drawPics(effect_list,sort_eff_list,eff_sampleDict,-1)
